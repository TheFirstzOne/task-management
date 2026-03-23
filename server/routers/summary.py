# -*- coding: utf-8 -*-
"""Summary router — Phase 21"""

import os
import tempfile

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session

from app.models.user import User
from app.services.task_service import TaskService
from app.services.time_tracking_service import TimeTrackingService
from server.deps import get_current_user, get_db

router = APIRouter()


@router.get("/time-by-task")
def time_by_task(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Return total tracked time grouped by task."""
    svc = TimeTrackingService(db)
    return svc.get_summary_by_task()


@router.get("/time-by-member")
def time_by_member(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Return total tracked time grouped by member."""
    svc = TimeTrackingService(db)
    return svc.get_summary_by_member()


@router.get("/export/excel")
def export_excel(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Export summary data as an Excel workbook with 3 sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl is not installed")

    task_svc = TaskService(db)
    time_svc = TimeTrackingService(db)

    # Collect data
    stats = task_svc.get_dashboard_stats()
    by_task = time_svc.get_summary_by_task()
    by_member = time_svc.get_summary_by_member()

    wb = openpyxl.Workbook()

    # ── Sheet 1: Dashboard Stats ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Dashboard Stats"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    ws1.append(["Metric", "Count"])
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    stat_rows = [
        ("Total Tasks", stats.get("total", 0)),
        ("Pending", stats.get("pending", 0)),
        ("In Progress", stats.get("in_progress", 0)),
        ("Review", stats.get("review", 0)),
        ("Done", stats.get("done", 0)),
        ("Cancelled", stats.get("cancelled", 0)),
        ("Overdue", stats.get("overdue", 0)),
    ]
    for row in stat_rows:
        ws1.append(row)

    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 12

    # ── Sheet 2: Time by Task ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Time by Task")
    ws2.append(["Task Title", "Total Minutes", "Hours"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in by_task:
        minutes = row.get("total_minutes", 0)
        ws2.append([row.get("title", ""), minutes, round(minutes / 60, 2)])

    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 12

    # ── Sheet 3: Time by Member ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Time by Member")
    ws3.append(["Member", "Total Minutes", "Hours", "Task Count"])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in by_member:
        ws3.append([
            row.get("name", ""),
            row.get("total_minutes", 0),
            round(row.get("hours", 0), 2),
            row.get("task_count", 0),
        ])

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 12

    # Save to temp file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
    except Exception as exc:
        if os.path.exists(path):
            os.unlink(path)
        raise HTTPException(status_code=500, detail=f"Excel export failed: {exc}")

    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="vindflow_summary.xlsx",
    )
