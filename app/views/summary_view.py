# -*- coding: utf-8 -*-
"""
SummaryView — Phase 5: Summary & Export UI
Features:
  - Overview stat cards  (total / done / in-progress / overdue / cancelled)
  - Status breakdown bar  (visual proportion per status)
  - Priority breakdown bar
  - Per-team table  (team | total | done | in-progress | overdue)
  - Per-member table  (name | team | assigned | done | overdue)
  - Filter: date range  (ช่วงวันที่)  +  team dropdown
  - Export to Excel  (openpyxl)
  - Export to PDF    (reportlab)
Flet 0.80.x — function-based
"""

from __future__ import annotations

import io
import os
import tempfile
from datetime import date, datetime, timedelta
from typing import List, Optional

import flet as ft
from sqlalchemy.orm import Session

from app.services.task_service import TaskService
from app.services.team_service import TeamService
from app.models.task import Task, TaskStatus, TaskPriority
from app.utils.theme import (
    BG_DARK, BG_CARD, BG_INPUT, BG_SIDEBAR,
    ACCENT, ACCENT2,
    TEXT_PRI, TEXT_SEC, BORDER,
    COLOR_PENDING, COLOR_IN_PROGRESS, COLOR_REVIEW,
    COLOR_DONE, COLOR_CANCELLED, COLOR_OVERDUE,
    COLOR_LOW, COLOR_MEDIUM, COLOR_HIGH, COLOR_URGENT,
    status_color, priority_color,
)

ALL_OPT = "ทั้งหมด"

# ── Status display order ───────────────────────────────────────────────────
_STATUS_ORDER = [
    TaskStatus.PENDING,
    TaskStatus.IN_PROGRESS,
    TaskStatus.REVIEW,
    TaskStatus.DONE,
    TaskStatus.CANCELLED,
]
_STATUS_COLORS = {
    TaskStatus.PENDING:     COLOR_PENDING,
    TaskStatus.IN_PROGRESS: COLOR_IN_PROGRESS,
    TaskStatus.REVIEW:      COLOR_REVIEW,
    TaskStatus.DONE:        COLOR_DONE,
    TaskStatus.CANCELLED:   COLOR_CANCELLED,
}
_PRIO_ORDER  = [TaskPriority.LOW, TaskPriority.MEDIUM,
                TaskPriority.HIGH, TaskPriority.URGENT]
_PRIO_COLORS = {
    TaskPriority.LOW:    COLOR_LOW,
    TaskPriority.MEDIUM: COLOR_MEDIUM,
    TaskPriority.HIGH:   COLOR_HIGH,
    TaskPriority.URGENT: COLOR_URGENT,
}


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN ENTRY
# ══════════════════════════════════════════════════════════════════════════════
def build_summary_view(db: Session, page: ft.Page) -> ft.Control:
    task_svc = TaskService(db)
    team_svc = TeamService(db)

    today = date.today()

    # ── State ──────────────────────────────────────────────────────
    state = {
        "filter_team":   ALL_OPT,
        "date_from":     None,   # date | None
        "date_to":       None,   # date | None
    }

    # ── Mutable content container ───────────────────────────────────
    content_col = ft.Column(spacing=20, expand=True, scroll=ft.ScrollMode.AUTO)

    # ── Export status snackbar helper ───────────────────────────────
    def _snack(msg: str, color: str = COLOR_DONE):
        page.snack_bar = ft.SnackBar(
            content=ft.Text(msg, color=TEXT_PRI),
            bgcolor=color,
        )
        page.snack_bar.open = True
        try:
            page.update()
        except Exception:
            pass

    # ══════════════════════════════════════════════════════════════
    #  DATA HELPERS
    # ══════════════════════════════════════════════════════════════
    def _get_filtered_tasks() -> List[Task]:
        tasks = task_svc.get_all_tasks()

        # Team filter
        if state["filter_team"] != ALL_OPT:
            try:
                tid = int(state["filter_team"])
                tasks = [t for t in tasks
                         if t.team_id == tid
                         or (t.assignee and t.assignee.team_id == tid)]
            except (ValueError, TypeError):
                pass

        # Date range filter (based on due_date)
        if state["date_from"]:
            dt_from = datetime.combine(state["date_from"], datetime.min.time())
            tasks = [t for t in tasks if t.due_date and t.due_date >= dt_from]
        if state["date_to"]:
            dt_to = datetime.combine(state["date_to"], datetime.max.time())
            tasks = [t for t in tasks if t.due_date and t.due_date <= dt_to]

        return tasks

    def _is_overdue(t: Task) -> bool:
        return (
            t.due_date is not None
            and t.due_date.date() < today
            and t.status not in (TaskStatus.DONE, TaskStatus.CANCELLED)
        )

    def _compute_stats(tasks: List[Task]) -> dict:
        total     = len(tasks)
        by_status = {s: sum(1 for t in tasks if t.status == s)
                     for s in _STATUS_ORDER}
        by_prio   = {p: sum(1 for t in tasks if t.priority == p)
                     for p in _PRIO_ORDER}
        overdue   = sum(1 for t in tasks if _is_overdue(t))
        return {
            "total":     total,
            "by_status": by_status,
            "by_prio":   by_prio,
            "overdue":   overdue,
        }

    def _per_team_rows(tasks: List[Task]):
        """Return list of (team_name, total, done, in_prog, overdue)."""
        teams = team_svc.get_all_teams()
        rows = []
        for team in teams:
            tt = [t for t in tasks if t.team_id == team.id]
            if not tt:
                continue
            done    = sum(1 for t in tt if t.status == TaskStatus.DONE)
            in_prog = sum(1 for t in tt if t.status == TaskStatus.IN_PROGRESS)
            ov      = sum(1 for t in tt if _is_overdue(t))
            rows.append((team.name, len(tt), done, in_prog, ov))
        # Unassigned team
        no_team = [t for t in tasks if t.team_id is None]
        if no_team:
            done    = sum(1 for t in no_team if t.status == TaskStatus.DONE)
            in_prog = sum(1 for t in no_team if t.status == TaskStatus.IN_PROGRESS)
            ov      = sum(1 for t in no_team if _is_overdue(t))
            rows.append(("(ไม่มีทีม)", len(no_team), done, in_prog, ov))
        return rows

    def _per_member_rows(tasks: List[Task]):
        """Return list of (name, team_name, total, done, overdue)."""
        users = task_svc.user_repo.get_all(active_only=False)
        rows = []
        for u in users:
            ut = [t for t in tasks if t.assignee_id == u.id]
            if not ut:
                continue
            done = sum(1 for t in ut if t.status == TaskStatus.DONE)
            ov   = sum(1 for t in ut if _is_overdue(t))
            team_name = u.team.name if u.team else "—"
            rows.append((u.name, team_name, len(ut), done, ov))
        return rows

    # ══════════════════════════════════════════════════════════════
    #  UI BUILDERS
    # ══════════════════════════════════════════════════════════════

    # ── Stat card ────────────────────────────────────────────────
    def _stat_card(label: str, value: int, color: str,
                   icon: str) -> ft.Container:
        return ft.Container(
            width=170,
            bgcolor=BG_CARD,
            border_radius=12,
            border=ft.border.all(1, BORDER),
            padding=ft.padding.all(18),
            content=ft.Column(
                controls=[
                    ft.Row(
                        controls=[
                            ft.Icon(icon, color=color, size=22),
                            ft.Text(label, size=12, color=TEXT_SEC),
                        ],
                        spacing=8,
                        vertical_alignment=ft.CrossAxisAlignment.CENTER,
                    ),
                    ft.Text(str(value), size=32,
                            weight=ft.FontWeight.BOLD, color=color),
                ],
                spacing=8,
            ),
        )

    # ── Horizontal proportion bar ─────────────────────────────────
    def _prop_bar(items: list[tuple[str, int, str]],
                  total: int) -> ft.Column:
        """
        items = [(label, count, color), ...]
        Builds a stacked bar + legend row.
        """
        bar_segments = []
        for label, count, col in items:
            if total > 0 and count > 0:
                bar_segments.append(
                    ft.Container(
                        expand=count,
                        height=14,
                        bgcolor=col,
                        tooltip=f"{label}: {count}",
                    )
                )

        bar = ft.Row(
            controls=bar_segments or [
                ft.Container(expand=1, height=14, bgcolor=BORDER)
            ],
            spacing=0,
            expand=True,
        )

        legend_items = []
        for label, count, col in items:
            legend_items.append(
                ft.Row(
                    controls=[
                        ft.Container(
                            width=10, height=10, border_radius=5, bgcolor=col,
                        ),
                        ft.Text(f"{label} ({count})", size=11, color=TEXT_SEC),
                    ],
                    spacing=6,
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                )
            )

        legend = ft.Row(controls=legend_items, spacing=16, wrap=True)

        return ft.Column(
            controls=[
                ft.Container(
                    content=bar,
                    border_radius=7,
                    clip_behavior=ft.ClipBehavior.ANTI_ALIAS,
                ),
                legend,
            ],
            spacing=8,
        )

    # ── Section header ────────────────────────────────────────────
    def _section(title: str, content: ft.Control) -> ft.Container:
        return ft.Container(
            bgcolor=BG_CARD,
            border_radius=12,
            border=ft.border.all(1, BORDER),
            padding=ft.padding.all(20),
            content=ft.Column(
                controls=[
                    ft.Text(title, size=15,
                            weight=ft.FontWeight.W_600, color=TEXT_PRI),
                    ft.Divider(height=1, color=BORDER),
                    content,
                ],
                spacing=12,
            ),
        )

    # ── Table helper ─────────────────────────────────────────────
    def _table_header(*cols: str) -> ft.Row:
        return ft.Row(
            controls=[
                ft.Container(
                    expand=True,
                    content=ft.Text(c, size=11, color=TEXT_SEC,
                                    weight=ft.FontWeight.W_600),
                )
                for c in cols
            ],
            spacing=4,
        )

    def _table_row(*vals, highlight_last_red: bool = False) -> ft.Container:
        cells = []
        for i, v in enumerate(vals):
            color = TEXT_PRI
            if highlight_last_red and i == len(vals) - 1 and int(v) > 0:
                color = COLOR_OVERDUE
            cells.append(
                ft.Container(
                    expand=True,
                    content=ft.Text(str(v), size=12, color=color),
                )
            )
        return ft.Container(
            content=ft.Row(controls=cells, spacing=4),
            padding=ft.padding.symmetric(vertical=6),
            border=ft.border.only(bottom=ft.BorderSide(1, BORDER + "66")),
        )

    # ══════════════════════════════════════════════════════════════
    #  EXPORT HELPERS
    # ══════════════════════════════════════════════════════════════
    def _export_excel(tasks: List[Task]):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            _snack("ไม่พบ openpyxl — กรุณา pip install openpyxl", COLOR_OVERDUE)
            return

        wb = openpyxl.Workbook()

        # ── Sheet 1: All tasks ─────────────────────────────────
        ws1 = wb.active
        ws1.title = "รายการงาน"

        hdr_fill = PatternFill("solid", fgColor="2563EB")   # ACCENT blue-white theme
        hdr_font = Font(color="FFFFFF", bold=True)

        headers = ["ID", "ชื่องาน", "สถานะ", "Priority",
                   "ผู้รับผิดชอบ", "ทีม", "วันกำหนด", "เกินกำหนด"]
        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        ws1.column_dimensions["B"].width = 35
        ws1.column_dimensions["C"].width = 14
        ws1.column_dimensions["D"].width = 12
        ws1.column_dimensions["E"].width = 18
        ws1.column_dimensions["F"].width = 16
        ws1.column_dimensions["G"].width = 16

        for row_idx, t in enumerate(tasks, 2):
            due_str  = t.due_date.strftime("%d/%m/%Y") if t.due_date else ""
            ov_str   = "ใช่" if _is_overdue(t) else ""
            assignee = t.assignee.name if t.assignee else ""
            team_nm  = t.team.name if t.team else ""
            row_data = [t.id, t.title, t.status.value, t.priority.value,
                        assignee, team_nm, due_str, ov_str]
            for col, val in enumerate(row_data, 1):
                ws1.cell(row=row_idx, column=col, value=val)

        # ── Sheet 2: Per-team summary ──────────────────────────
        ws2 = wb.create_sheet("สรุปตามทีม")
        team_rows = _per_team_rows(tasks)
        h2 = ["ทีม", "งานทั้งหมด", "เสร็จแล้ว", "กำลังทำ", "เกินกำหนด"]
        for col, h in enumerate(h2, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
        for ri, r in enumerate(team_rows, 2):
            for ci, v in enumerate(r, 1):
                ws2.cell(row=ri, column=ci, value=v)

        # ── Sheet 3: Per-member summary ────────────────────────
        ws3 = wb.create_sheet("สรุปตามสมาชิก")
        member_rows = _per_member_rows(tasks)
        h3 = ["ชื่อ", "ทีม", "งานทั้งหมด", "เสร็จแล้ว", "เกินกำหนด"]
        for col, h in enumerate(h3, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
        for ri, r in enumerate(member_rows, 2):
            for ci, v in enumerate(r, 1):
                ws3.cell(row=ri, column=ci, value=v)

        # ── Save ───────────────────────────────────────────────
        ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = os.path.join(os.path.expanduser("~"), "Desktop",
                           f"TaskReport_{ts}.xlsx")
        wb.save(out)
        _snack(f"บันทึก Excel แล้ว: {os.path.basename(out)}")

    def _export_pdf(tasks: List[Task]):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors as rl_colors
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle,
                Paragraph, Spacer,
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            _snack("ไม่พบ reportlab — กรุณา pip install reportlab", COLOR_OVERDUE)
            return

        # ── Register Thai font (Tahoma — built-in Windows, supports Thai) ──
        import sys
        _FONT_NAME = "Tahoma"
        _font_registered = False
        _font_candidates = [
            r"C:\Windows\Fonts\tahoma.ttf",
            "/usr/share/fonts/truetype/msttcorefonts/Tahoma.ttf",
        ]
        for _fp in _font_candidates:
            if os.path.exists(_fp):
                try:
                    pdfmetrics.registerFont(TTFont(_FONT_NAME, _fp))
                    _font_registered = True
                except Exception:
                    pass
                break
        if not _font_registered:
            _FONT_NAME = "Helvetica"   # fallback — no Thai support

        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        out  = os.path.join(os.path.expanduser("~"), "Desktop",
                            f"TaskReport_{ts}.pdf")

        doc = SimpleDocTemplate(
            out, pagesize=A4,
            leftMargin=2*cm, rightMargin=2*cm,
            topMargin=2*cm, bottomMargin=2*cm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "title", parent=styles["Heading1"],
            fontSize=16, textColor=rl_colors.HexColor("#2563EB"),
            fontName=_FONT_NAME,
        )
        h2_style = ParagraphStyle(
            "h2", parent=styles["Heading2"],
            fontSize=12, textColor=rl_colors.HexColor("#FFFFFF"),
            backColor=rl_colors.HexColor("#2563EB"),
            spaceBefore=14,
            fontName=_FONT_NAME,
        )
        normal = ParagraphStyle(
            "normal", parent=styles["Normal"],
            fontSize=9, textColor=rl_colors.HexColor("#1E293B"),
            fontName=_FONT_NAME,
        )

        _HDR_BG   = rl_colors.HexColor("#2563EB")   # ACCENT
        _ACCENT   = rl_colors.HexColor("#FFFFFF")   # white text on blue header
        _TEXT     = rl_colors.HexColor("#1E293B")   # TEXT_PRI
        _TEXT_SEC = rl_colors.HexColor("#64748B")   # TEXT_SEC
        _RED      = rl_colors.HexColor("#EF4444")   # COLOR_OVERDUE
        _BORDER   = rl_colors.HexColor("#CBD5E1")   # BORDER

        def _tbl_style(ncols: int) -> TableStyle:
            return TableStyle([
                ("BACKGROUND",  (0, 0), (ncols-1, 0), _HDR_BG),
                ("TEXTCOLOR",   (0, 0), (ncols-1, 0), _ACCENT),
                ("FONTNAME",    (0, 0), (ncols-1, 0), _FONT_NAME),   # header font
                ("FONTSIZE",    (0, 0), (ncols-1, 0), 9),
                ("TEXTCOLOR",   (0, 1), (ncols-1, -1), _TEXT),
                ("FONTNAME",    (0, 1), (ncols-1, -1), _FONT_NAME),   # data font
                ("FONTSIZE",    (0, 1), (ncols-1, -1), 8),
                ("GRID",        (0, 0), (ncols-1, -1), 0.4, _BORDER),
                ("ROWBACKGROUNDS", (0, 1), (ncols-1, -1),
                 [rl_colors.HexColor("#FFFFFF"), rl_colors.HexColor("#F0F4F8")]),
                ("VALIGN",      (0, 0), (ncols-1, -1), "MIDDLE"),
                ("TOPPADDING",  (0, 0), (ncols-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (ncols-1, -1), 5),
            ])

        story = []

        # Title
        gen_time = datetime.now().strftime("%d/%m/%Y %H:%M")
        story.append(Paragraph("รายงานสรุปงาน (Task Report)", title_style))
        story.append(Paragraph(f"สร้างเมื่อ: {gen_time}", normal))
        story.append(Spacer(1, 0.4*cm))

        # Stats summary
        stats = _compute_stats(tasks)
        story.append(Paragraph("ภาพรวม", h2_style))
        overview_data = [
            ["งานทั้งหมด", "เสร็จแล้ว", "กำลังทำ", "รอดำเนินการ", "เกินกำหนด"],
            [
                str(stats["total"]),
                str(stats["by_status"].get(TaskStatus.DONE, 0)),
                str(stats["by_status"].get(TaskStatus.IN_PROGRESS, 0)),
                str(stats["by_status"].get(TaskStatus.PENDING, 0)),
                str(stats["overdue"]),
            ],
        ]
        t1 = Table(overview_data, colWidths=[3*cm]*5)
        t1.setStyle(_tbl_style(5))
        story.append(t1)
        story.append(Spacer(1, 0.4*cm))

        # Per-team table
        team_rows = _per_team_rows(tasks)
        if team_rows:
            story.append(Paragraph("สรุปตามทีม", h2_style))
            tdata = [["ทีม", "งานทั้งหมด", "เสร็จแล้ว", "กำลังทำ", "เกินกำหนด"]]
            for r in team_rows:
                tdata.append(list(r))
            t2 = Table(tdata, colWidths=[5*cm, 3*cm, 3*cm, 3*cm, 3*cm])
            style2 = _tbl_style(5)
            # Colour overdue column red if > 0
            for ri, r in enumerate(tdata[1:], 1):
                if int(r[4]) > 0:
                    style2.add("TEXTCOLOR", (4, ri), (4, ri), _RED)
            t2.setStyle(style2)
            story.append(t2)
            story.append(Spacer(1, 0.4*cm))

        # Per-member table
        member_rows = _per_member_rows(tasks)
        if member_rows:
            story.append(Paragraph("สรุปตามสมาชิก", h2_style))
            mdata = [["ชื่อ", "ทีม", "งานทั้งหมด", "เสร็จแล้ว", "เกินกำหนด"]]
            for r in member_rows:
                mdata.append(list(r))
            t3 = Table(mdata, colWidths=[4*cm, 4*cm, 3*cm, 3*cm, 3*cm])
            style3 = _tbl_style(5)
            for ri, r in enumerate(mdata[1:], 1):
                if int(r[4]) > 0:
                    style3.add("TEXTCOLOR", (4, ri), (4, ri), _RED)
            t3.setStyle(style3)
            story.append(t3)
            story.append(Spacer(1, 0.4*cm))

        # All tasks table
        story.append(Paragraph("รายการงานทั้งหมด", h2_style))
        cols = ["ID", "ชื่องาน", "สถานะ", "Priority", "ผู้รับผิดชอบ", "วันกำหนด"]
        task_data = [cols]
        for t in tasks:
            due_str  = t.due_date.strftime("%d/%m/%Y") if t.due_date else "—"
            assignee = t.assignee.name if t.assignee else "—"
            task_data.append([
                str(t.id), t.title[:40], t.status.value,
                t.priority.value, assignee, due_str,
            ])
        t4 = Table(task_data,
                   colWidths=[1*cm, 6*cm, 3*cm, 2.5*cm, 3*cm, 2.5*cm])
        style4 = _tbl_style(6)
        t4.setStyle(style4)
        story.append(t4)

        doc.build(story)
        _snack(f"บันทึก PDF แล้ว: {os.path.basename(out)}")

    # ══════════════════════════════════════════════════════════════
    #  REBUILD CONTENT
    # ══════════════════════════════════════════════════════════════
    def _rebuild():
        tasks = _get_filtered_tasks()
        stats = _compute_stats(tasks)
        total = stats["total"]

        # ── Stat cards row ──────────────────────────────────────
        cards = ft.Row(
            controls=[
                _stat_card("งานทั้งหมด", total,
                           ACCENT, ft.Icons.ASSIGNMENT_OUTLINED),
                _stat_card("เสร็จแล้ว",
                           stats["by_status"][TaskStatus.DONE],
                           COLOR_DONE, ft.Icons.CHECK_CIRCLE_OUTLINE),
                _stat_card("กำลังทำ",
                           stats["by_status"][TaskStatus.IN_PROGRESS],
                           COLOR_IN_PROGRESS, ft.Icons.TIMELAPSE),
                _stat_card("รอดำเนินการ",
                           stats["by_status"][TaskStatus.PENDING],
                           COLOR_PENDING, ft.Icons.PENDING_OUTLINED),
                _stat_card("เกินกำหนด", stats["overdue"],
                           COLOR_OVERDUE, ft.Icons.WARNING_AMBER_OUTLINED),
            ],
            spacing=12,
            wrap=True,
            vertical_alignment=ft.CrossAxisAlignment.START,
        )

        # ── Status bar ──────────────────────────────────────────
        status_items = [
            (s.value,
             stats["by_status"][s],
             _STATUS_COLORS[s])
            for s in _STATUS_ORDER
            if stats["by_status"][s] > 0
        ]
        status_section = _section(
            "สัดส่วนตามสถานะ",
            _prop_bar(status_items, total),
        )

        # ── Priority bar ────────────────────────────────────────
        prio_items = [
            (p.value,
             stats["by_prio"][p],
             _PRIO_COLORS[p])
            for p in _PRIO_ORDER
            if stats["by_prio"][p] > 0
        ]
        prio_section = _section(
            "สัดส่วนตาม Priority",
            _prop_bar(prio_items, total),
        )

        # ── Per-team table ──────────────────────────────────────
        team_rows = _per_team_rows(tasks)
        team_rows_ctrl = [
            _table_header("ทีม", "งานทั้งหมด", "เสร็จแล้ว", "กำลังทำ", "เกินกำหนด"),
        ]
        if team_rows:
            for r in team_rows:
                team_rows_ctrl.append(
                    _table_row(*r, highlight_last_red=True)
                )
        else:
            team_rows_ctrl.append(
                ft.Text("ไม่มีข้อมูล", size=12, color=TEXT_SEC)
            )
        team_section = _section(
            "สรุปตามทีม",
            ft.Column(controls=team_rows_ctrl, spacing=0),
        )

        # ── Per-member table ────────────────────────────────────
        member_rows = _per_member_rows(tasks)
        member_rows_ctrl = [
            _table_header("ชื่อ", "ทีม", "งานทั้งหมด", "เสร็จแล้ว", "เกินกำหนด"),
        ]
        if member_rows:
            for r in member_rows:
                member_rows_ctrl.append(
                    _table_row(*r, highlight_last_red=True)
                )
        else:
            member_rows_ctrl.append(
                ft.Text("ไม่มีข้อมูล", size=12, color=TEXT_SEC)
            )
        member_section = _section(
            "สรุปตามสมาชิก",
            ft.Column(controls=member_rows_ctrl, spacing=0),
        )

        content_col.controls = [
            cards,
            ft.Row(controls=[status_section, prio_section],
                   spacing=16, expand=True),
            team_section,
            member_section,
        ]
        try:
            content_col.update()
        except Exception:
            pass

    # ══════════════════════════════════════════════════════════════
    #  FILTER BAR
    # ══════════════════════════════════════════════════════════════
    def _team_options():
        teams = team_svc.get_all_teams()
        return [ft.dropdown.Option(ALL_OPT)] + [
            ft.dropdown.Option(key=str(t.id), text=t.name) for t in teams
        ]

    dd_team = ft.Dropdown(
        label="ทีม", width=150,
        options=_team_options(), value=ALL_OPT,
        border_color=BORDER, focused_border_color=ACCENT,
        label_style=ft.TextStyle(color=TEXT_SEC),
        color=TEXT_PRI, bgcolor=BG_INPUT, border_radius=8,
        text_size=12,
        on_select=lambda e: _on_filter("filter_team", e.control.value),
    )

    # Date from / to as simple text fields (dd/mm/yyyy)
    tf_date_from = ft.TextField(
        label="วันที่เริ่ม (dd/mm/yyyy)", width=170,
        hint_text="01/01/2568",
        border_color=BORDER, focused_border_color=ACCENT,
        label_style=ft.TextStyle(color=TEXT_SEC, size=11),
        color=TEXT_PRI, bgcolor=BG_INPUT, border_radius=8,
        text_size=12,
        on_submit=lambda e: _apply_dates(),
        on_blur=lambda e: _apply_dates(),
    )
    tf_date_to = ft.TextField(
        label="วันที่สิ้นสุด (dd/mm/yyyy)", width=170,
        hint_text="31/12/2568",
        border_color=BORDER, focused_border_color=ACCENT,
        label_style=ft.TextStyle(color=TEXT_SEC, size=11),
        color=TEXT_PRI, bgcolor=BG_INPUT, border_radius=8,
        text_size=12,
        on_submit=lambda e: _apply_dates(),
        on_blur=lambda e: _apply_dates(),
    )

    def _parse_date_field(val: str) -> Optional[date]:
        val = val.strip()
        if not val:
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
        # Try Buddhist Era → CE
        try:
            d, m, y_be = val.split("/")
            return date(int(y_be) - 543, int(m), int(d))
        except Exception:
            return None

    def _apply_dates():
        state["date_from"] = _parse_date_field(tf_date_from.value or "")
        state["date_to"]   = _parse_date_field(tf_date_to.value or "")
        _rebuild()

    def _on_filter(key: str, val):
        state[key] = val
        _rebuild()

    def _clear_filters(e):
        state["filter_team"] = ALL_OPT
        state["date_from"]   = None
        state["date_to"]     = None
        dd_team.value        = ALL_OPT
        tf_date_from.value   = ""
        tf_date_to.value     = ""
        try:
            dd_team.update()
        except Exception:
            pass
        try:
            tf_date_from.update()
        except Exception:
            pass
        try:
            tf_date_to.update()
        except Exception:
            pass
        _rebuild()

    def _on_export_excel(e):
        _export_excel(_get_filtered_tasks())

    def _on_export_pdf(e):
        _export_pdf(_get_filtered_tasks())

    # ── Refresh button ────────────────────────────────────────────
    def _on_refresh(e):
        _rebuild()

    # ══════════════════════════════════════════════════════════════
    #  TOP BAR
    # ══════════════════════════════════════════════════════════════
    top_bar = ft.Row(
        controls=[
            ft.Column(
                controls=[
                    ft.Text("สรุปงาน", size=24,
                            weight=ft.FontWeight.BOLD, color=TEXT_PRI),
                    ft.Text("ภาพรวมและสถิติของงานทั้งหมด",
                            size=13, color=TEXT_SEC),
                ],
                spacing=2,
            ),
            ft.Row(
                controls=[
                    ft.Button(
                        content=ft.Text("Export Excel"),
                        icon=ft.Icons.TABLE_CHART_OUTLINED,
                        bgcolor=BG_CARD,
                        color=COLOR_DONE,
                        style=ft.ButtonStyle(
                            side=ft.BorderSide(1, COLOR_DONE + "66"),
                            shape=ft.RoundedRectangleBorder(radius=8),
                        ),
                        on_click=_on_export_excel,
                    ),
                    ft.Button(
                        content=ft.Text("Export PDF"),
                        icon=ft.Icons.PICTURE_AS_PDF_OUTLINED,
                        bgcolor=BG_CARD,
                        color=COLOR_OVERDUE,
                        style=ft.ButtonStyle(
                            side=ft.BorderSide(1, COLOR_OVERDUE + "66"),
                            shape=ft.RoundedRectangleBorder(radius=8),
                        ),
                        on_click=_on_export_pdf,
                    ),
                    ft.IconButton(
                        icon=ft.Icons.REFRESH,
                        icon_color=TEXT_SEC, icon_size=20,
                        tooltip="รีเฟรช",
                        on_click=_on_refresh,
                    ),
                ],
                spacing=8,
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
            ),
        ],
        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
        vertical_alignment=ft.CrossAxisAlignment.START,
    )

    filter_bar = ft.Row(
        controls=[
            ft.Icon(ft.Icons.FILTER_LIST, color=TEXT_SEC, size=18),
            ft.Text("กรองข้อมูล:", size=12, color=TEXT_SEC),
            dd_team,
            tf_date_from,
            tf_date_to,
            ft.TextButton(
                "ล้าง",
                style=ft.ButtonStyle(color=TEXT_SEC),
                on_click=_clear_filters,
            ),
        ],
        spacing=10,
        vertical_alignment=ft.CrossAxisAlignment.CENTER,
        wrap=True,
    )

    # ── Initial render ────────────────────────────────────────────
    _rebuild()

    return ft.Container(
        expand=True,
        bgcolor=BG_DARK,
        padding=ft.padding.all(24),
        content=ft.Column(
            controls=[
                top_bar,
                ft.Divider(height=12, color=BORDER),
                filter_bar,
                ft.Divider(height=12, color=BORDER),
                content_col,
            ],
            spacing=0,
            expand=True,
        ),
    )